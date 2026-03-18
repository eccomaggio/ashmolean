from pathlib import Path
import argparse
import logging
from csv import reader, writer
from openpyxl import load_workbook, Workbook  # type: ignore[import]
from openpyxl.utils import get_column_letter
import json
from datetime import datetime


def argument_parser() -> tuple[Path, Path, Path]:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Process some text files.")
    parser.add_argument(
        "-s", "--source", type=Path, default=Path("input.txt"), help="Source file path"
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("output.csv"),
        help="Destination file path",
    )
    parser.add_argument(
        "-c",
        "--concordance",
        type=Path,
        default=Path("concordance.xlsx"),
        help="Concordance file path",
    )
    args = parser.parse_args()
    return (args.source, args.output, args.concordance)


def read_lines(file_path: Path) -> list[str]:
    """Read lines from a file and return them as a list."""
    with file_path.open("r", encoding="utf-8") as file:
        raw_lines = file.readlines()
    # if ord(raw_lines[0][0]) == 65279:
    #     raw_lines[0] = raw_lines[0][1:]
    raw_lines[0] = remove_bom(raw_lines[0])
    raw_lines = [line.strip() for line in raw_lines]
    return raw_lines
    # return file.readlines()


def remove_bom(line: str) -> str:
    # print(f">>> start character={ord(line[0])} ({ord(line[0]) == 65279})")
    if ord(line[0]) == 65279:
        # line = line[2:]
        line = line[1:]
        logging.info(f"Removed BOM from start of file. <{line[:10]}>")
    else:
        logging.error("No BOM found.")
    return line


def write_lines(file_path: Path, lines: list[str]) -> None:
    """Write a list of lines to a file."""
    with file_path.open("w", encoding="utf-8") as file:
        file.writelines(lines)


def write_csv(file_path: Path, data: list) -> None:
    """Write a list of lists to a CSV file."""
    with file_path.open("w", encoding="utf-8", newline="") as file:
        csv_writer = writer(file)
        csv_writer.writerows(data)


def read_csv(file_path: Path) -> list[list[str]]:
    """Read a CSV file and return its content as a list of lists."""
    with file_path.open("r", encoding="utf-8") as file:
        csv_reader = reader(file)
        return list(csv_reader)


def load_json_file(filename) -> dict:
    """
    Reads a JSON file and returns a Python dictionary.
    """
    file_path = Path(filename)
    if not file_path.exists():
        print(f"Error: {filename} not found.")
        return {}
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    logging.info(f"Successfully read in {filename}")
    return data


def export_dict_to_json(data:dict[str, list[str]], file_path:Path):
    """
    Saves a Python dictionary to a JSON file with human-readable formatting.
    """
    filename = file_path.with_suffix(".json")
    with open(filename, 'w', encoding='utf-8') as f:
        # indent=4 makes the file readable (otherwise it's one long line)
        # ensure_ascii=False allows non-English characters to save correctly
        json.dump(data, f, indent=4, ensure_ascii=False)
    logging.info(f"Successfully saved to {filename}")


def export_to_excel(file_path:Path, data_rows:list) -> None:
    """
    Takes a list of tuples and saves them to an Excel file.

    Args:
        data_rows (list[tuple]): The data including the header row.
        filename (str): The desired output file name.
    """
    format_cols_as_dates = [3]
    filename = file_path.with_suffix(".xlsx").absolute()
    wb = Workbook()
    ws = wb.active
    headers = data_rows[0]
    body_rows = data_rows[1:]
    ws.append(headers)
    for row_num, row in enumerate(body_rows):
        # Convert the tuple to a list so we can modify the 3rd column (index 2)
        row_list = list(row)
        for col_index in format_cols_as_dates:
            try:
                # Coerce index 2 (Column 3) to a datetime object
                # Adjust '%Y-%m-%d' to match your specific string format
                date_string = row_list[col_index]
                # row_list[col_index] = datetime.strptime(date_string, '%Y-%m-%d')
                row_list[col_index] = datetime.strptime(date_string, '%d-%m-%Y')
            except (ValueError, TypeError):
                logging.warning(f"Row {row_num}: The value ({date_string}) cannot be expressed as a date.")
        ws.append(row_list)

    # Optional: Apply a specific date format to Column C (index 3)
    # This ensures Excel displays it exactly how you want
    # for cell in ws['D']:
    for col_index in format_cols_as_dates:
        letter = get_column_letter(col_index + 1)
        for cell in ws[letter]:
            if cell.row > 1: # Skip header
                cell.number_format = 'dd-mm-yyyy'
    wb.save(filename)
    logging.info(f"File saved successfully as {filename}")


def extract_from_excel(excel_file_address: Path) -> list[list[str]]:
    """
    excel seems pretty random in how it assigns string/int/float, so...
    this routine coerces everything into a string,
    strips ".0" from misrecognised floats
    & removes trailing spaces
    """
    if error:= is_file_locked(excel_file_address):
        logging.critical(f"The concordance file {error}.")
    excel_file_name: str = str(excel_file_address.resolve())
    excel_sheet = load_workbook(filename=excel_file_name).active
    logging.info(f"Opening concordance file {excel_file_name}...")
    sheet = []
    if excel_sheet:
        for i, excel_row in enumerate(excel_sheet.iter_rows(min_row=2, values_only=True)):
            # print(f"DEBUG: {i}")
            row = []
            for col in excel_row:
                if col:
                    data = str(col).strip()
                    data = trim_mistaken_decimals(data)
                else:
                    data = ""
                row.append(data)
            sheet.append(row)
    else:
        logging.critical(f"No worksheet found for {excel_file_name}")
    return sheet
# def extract_from_excel(excel_file_address: Path) -> list[list[str]]:
#     """
#     excel seems pretty random in how it assigns string/int/float, so...
#     this routine coerces everything into a string,
#     strips ".0" from misrecognised floats
#     & removes trailing spaces
#     """
#     excel_file_name: str = str(excel_file_address.resolve())
#     excel_sheet = load_workbook(filename=excel_file_name).active
#     sheet = []
#     if excel_sheet:
#         for excel_row in excel_sheet.iter_rows(min_row=2, values_only=True):
#             row = []
#             if not excel_row[0]:
#                 break
#             for col in excel_row:
#                 if col:
#                     data = str(col).strip()
#                     data = trim_mistaken_decimals(data)
#                 else:
#                     data = ""
#                 row.append(data)
#             sheet.append(row)
#     return sheet


def is_file_locked(filepath: str | Path) -> str:
    path = Path(filepath)
    if not path.exists():
        return "does not exist"
    try:
        # Attempting to rename the file to its current name
        # If open in another app, Windows/MacOS will raise an OSError
        path.rename(path)
        return "is locked"
    except OSError:
        return ""


def check_for_empty_line(row: list[str]) -> bool:
    line_is_empty = "".join(row).strip() == ""
    return line_is_empty


def trim_mistaken_decimals(string: str) -> str:
    if string.endswith(".0"):
        string = string[:-2]
    return string
