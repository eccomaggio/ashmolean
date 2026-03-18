"""
Input:
./[concordance file name]
./text_files/[marked up text files]

Output:
./csv_files/[file for upload to Museum+]
The app will first create a concordance file linking section number to museum+ object id & catalogue number
It will then use this to create a .csv file composed of sections as literature to be uploaded to Museum+ with the correct object id

Markup:
[One per line:]

@@META:PUB_DATE=DD/MM/YYYY
the publication date of the work

@@NEW:1 or 1&2&...
the section title or titles

@@PROCESS
Ignore any lines before this after encountering a @@NEW

[Inline commands:]

@@LINK: ... @@ -> e.g. @@LINK:No. 20@@
insert link text -- but currently not used. The markup is simply discarded, leaving the bare text inside untouched.


"""

from pathlib import Path
from csv import reader, writer
import argparse
from openpyxl import load_workbook, Workbook  # type: ignore[import]
import json

from dataclasses import dataclass, field
from datetime import datetime

import logging
import sys
import re
from collections import defaultdict

from pprint import pprint

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s:%(message)s",
    handlers=[
        # Configure filemode and encoding here instead
        logging.FileHandler("textExtract.log", mode="w", encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ],
)


@dataclass
class Overview:
    count: defaultdict[str, int] = field(default_factory=lambda: defaultdict(int))
    missing: defaultdict[str, list[int]] = field(
        default_factory=lambda: defaultdict(list)
    )

overview = Overview()


class Content:
    def __init__(self, pd: str, ps: dict, cl: list, line: str, csk: list, cinfo: dict, ci=True):
        self.pub_date: str = pd
        self.processed_sections: dict[str, list[str]] = ps
        self.current_lines: list[str] = cl
        self.line: str = line
        self.current_section_keys: list[str] = csk
        self.current_info: dict[str, str] = cinfo
        self.currently_ignoring: bool = ci

    def update_processed_sections(self):
        for key in self.current_section_keys:
            self.processed_sections[key] = self.current_lines
        self.current_lines = []
        self.current_section_keys = []

    def start_new_section(self, new_section_keys: list[str]):
        self.update_processed_sections()
        self.current_section_keys = new_section_keys
        self.currently_ignoring = True

    def update_current_lines(self):
        if self.line:
            self.current_lines.append(self.line)
            self.line = ""

    def add_to_line(self, part:str):
        # print(f">>>> {part[:50]}, {self.currently_ignoring}")
        if not self.currently_ignoring:
            self.line += part


# class Instruction(Enum):
#     NONE = auto()
#     NEW_SECTION = auto()
#     PROCESS = auto()
#     IGNORE = auto()
#     META = auto()
#     UNKNOWN = auto()


# @dataclass
# class Command:
#     order: Instruction
#     details: list[str]
@dataclass
class Command:
    verb: str
    object_list: list[str]


type ExcelRow = tuple[str, str, str, str, str, str, str, str, str, str, str, str, str]


def argument_parser() -> tuple[Path, Path, Path]:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Process some text files.")
    parser.add_argument(
        "-s", "--source", type=Path, default=Path("input.txt"), help="Source file path"
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("output.csv"),
        help="Destination file path",
    )
    parser.add_argument(
        "-c",
        "--concordance",
        type=Path,
        default=Path("concordance.xlsx"),
        help="Concordance file path",
    )
    args = parser.parse_args()
    return (args.source, args.output, args.concordance)


def is_file_locked(filepath: str | Path) -> str:
    path = Path(filepath)
    if not path.exists():
        return "does not exist"
    try:
        # Attempting to rename the file to its current name
        # If open in another app, Windows/MacOS will raise an OSError
        path.rename(path)
        return "is locked"
    except OSError:
        return ""


def read_lines(file_path: Path) -> list[str]:
    """Read lines from a file and return them as a list."""
    with file_path.open("r", encoding="utf-8") as file:
        raw_lines = file.readlines()
    # if ord(raw_lines[0][0]) == 65279:
    #     raw_lines[0] = raw_lines[0][1:]
    raw_lines[0] = remove_bom(raw_lines[0])
    raw_lines = [line.strip() for line in raw_lines]
    return raw_lines
    # return file.readlines()


def remove_bom(line: str) -> str:
    # print(f">>> start character={ord(line[0])} ({ord(line[0]) == 65279})")
    if ord(line[0]) == 65279:
        # line = line[2:]
        line = line[1:]
        logging.info(f"Removed BOM from start of file. <{line[:10]}>")
    else:
        logging.error("No BOM found.")
    return line


def write_lines(file_path: Path, lines: list[str]) -> None:
    """Write a list of lines to a file."""
    with file_path.open("w", encoding="utf-8") as file:
        file.writelines(lines)


def export_to_csv(file_path: Path, data_rows: list) -> None:
    """Write a list of lists to a CSV file."""
    with file_path.open("w", encoding="utf-8", newline="") as file:
        csv_writer = writer(file)
        csv_writer.writerows(data_rows)


# def export_to_excel(data_rows, filename="output.xlsx"):
def export_to_excel(file_path:Path, data_rows:list):
    """
    Takes a list of tuples and saves them to an Excel file.

    Args:
        data_rows (list[tuple]): The data including the header row.
        filename (str): The desired output file name.
    """
    # # Create a new workbook and select the active sheet
    # wb = Workbook()
    # ws = wb.active
    # ws.title = "Data Export"
    # filename = file_path.with_suffix(".xlsx").absolute()
    # for row in data_rows:
    #     ws.append(row)
    # wb.save(filename)
    # logging.info(f"File saved successfully as {filename}")
    filename = file_path.with_suffix(".xlsx").absolute()
    wb = Workbook()
    ws = wb.active
    headers = data_rows[0]
    body_rows = data_rows[1:]
    ws.append(headers)
    for row in body_rows:
        # Convert the tuple to a list so we can modify the 3rd column (index 2)
        row_list = list(row)
        try:
            # Coerce index 2 (Column 3) to a datetime object
            # Adjust '%Y-%m-%d' to match your specific string format
            date_string = row_list[2]
            row_list[3] = datetime.strptime(date_string, '%Y-%m-%d')
        except (ValueError, TypeError):
            # Fallback: if conversion fails, keep the original string
            # or handle the error as needed
            pass
        ws.append(row_list)

    # Optional: Apply a specific date format to Column C (index 3)
    # This ensures Excel displays it exactly how you want
    for cell in ws['D']:
        if cell.row > 1: # Skip header
            # cell.number_format = 'yyyy-mm-dd'
            cell.number_format = 'dd-mm-yyyy'
    wb.save(filename)
    logging.info(f"File saved successfully as {filename}")


def read_csv(file_path: Path) -> list[list[str]]:
    """Read a CSV file and return its content as a list of lists."""
    with file_path.open("r", encoding="utf-8") as file:
        csv_reader = reader(file)
        return list(csv_reader)


def load_json_file(filename):
    """
    Reads a JSON file and returns a Python dictionary.
    """
    file_path = Path(filename)
    if not file_path.exists():
        print(f"Error: {filename} not found.")
        return None
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    logging.info(f"Successfully read in {filename}")
    return data


def export_dict_to_json(data:dict[str, list[str]], file_path:Path):
    """
    Saves a Python dictionary to a JSON file with human-readable formatting.
    """
    filename = file_path.with_suffix(".json")
    with open(filename, 'w', encoding='utf-8') as f:
        # indent=4 makes the file readable (otherwise it's one long line)
        # ensure_ascii=False allows non-English characters to save correctly
        json.dump(data, f, indent=4, ensure_ascii=False)
    logging.info(f"Successfully saved to {filename}")


def extract_from_excel(excel_file_address: Path) -> list[list[str]]:
    """
    excel seems pretty random in how it assigns string/int/float, so...
    this routine coerces everything into a string,
    strips ".0" from misrecognised floats
    & removes trailing spaces
    """
    if error:= is_file_locked(excel_file_address):
        logging.critical(f"The concordance file {error}.")
    excel_file_name: str = str(excel_file_address.resolve())
    excel_sheet = load_workbook(filename=excel_file_name).active
    logging.info(f"Opening concordance file {excel_file_name}...")
    sheet = []
    if excel_sheet:
        for i, excel_row in enumerate(excel_sheet.iter_rows(min_row=2, values_only=True)):
            # print(f"DEBUG: {i}")
            row = []
            for col in excel_row:
                if col:
                    data = str(col).strip()
                    data = trim_mistaken_decimals(data)
                else:
                    data = ""
                row.append(data)
            sheet.append(row)
    else:
        logging.critical(f"No worksheet found for {excel_file_name}")
    return sheet


def check_for_empty_line(row:list[str]) -> bool:
    line_is_empty = "".join(row).strip() == ""
    return line_is_empty


def trim_mistaken_decimals(string: str) -> str:
    if string.endswith(".0"):
        string = string[:-2]
    return string


def make_concordance_from_excel(excel_file_path:Path) -> dict[str, list[str]]:
    """
    concordance layout with example data:
    row 0*  1*                   2                           3                   4                       5 (key)        6 notes
    ObjID	ObjectNumber	    ObjectNumberSorted	        ReferenceNumber	    ObjectTitle	            PW:litCatNo
    742939	WA1947.191.176.1	WA1947.00191.00176.00001	Penny (1992) 1	    Flagellator of Christ	1

    PW:litCatNo (column 5) is the number given in the publication, so it is the key to the object id
    It takes the first worksheet regardless of name.
    It returns a dictionary: { section id : [museum+ object number, catalogue number]}
    """
    raw = extract_from_excel(excel_file_path)
    # pprint(raw)
    concordance = normalise_concordance(raw)
    # pprint(concordance)
    return concordance


def normalise_concordance(raw: list[list[str]]) -> dict[str, list[str]]:
    concordance: dict[str, list[str]] = {}
    for row in raw:
        object_id = row[0]
        object_num = row[1]
        if row[5].isnumeric():
            concordance[row[5]] = [object_id, object_num]
    return concordance


def process(raw_lines: list[str], concordance: dict[str, list[str]]) -> Content:
    """
    This will replace group_lines(), get_command(), get_inline_commands() & parse_inline_command()
    This sorts the lines (paragraphs) into sections
    filters out lines that aren't needed
    and expands any inline commands
    """
    content: Content = Content("", {}, [], "", [],{})
    for i, raw_line in enumerate(raw_lines):
        if not raw_line:
            continue
        phrases = chunk_by_command(raw_line)
        for _type, cmd, other in phrases:
            match _type:
                case "none":
                    content.add_to_line(other)
                case "V":
                    content = process_verb(cmd, content)
                case "VO":
                    content = process_verb_object(cmd, content, concordance)
                case "unknown":
                    msg = f">>>> UNKNOWN command: {cmd}"
                    logging.error(msg)
        content.update_current_lines()
    if content.current_lines:
        content.update_processed_sections()
    if content.pub_date:
        logging.info(f"Pub date: {content.pub_date}")
    else:
        logging.critical("No publication date found!")
    return content


def chunk_by_command(line: str) -> list[tuple[str, Command, str]]:
    """
    Given a string with embedded commands, marked by @@
    returns a list conti
    """
    open_cmd = False
    output = []
    pre_parsed_line = [el for el in re.split(r"(@@)", line) if el]
    ## re group preserves "@@": allows it to be matched with command to distinguish "PROCESS" (text) vs "@@PROCESS" (cmd)
    for phrase in pre_parsed_line:
        if phrase == "@@":
            open_cmd = not open_cmd
            continue
        if open_cmd:
            _type, cmd, other = parse_command(phrase)
        else:
            _type = "none"
            cmd = Command("", [])
            other = phrase
            # print(f">>>> {phrase[:50]}")
        output.append((_type, cmd, other))
    return output


def parse_command(phrase: str) -> tuple[str, Command, str]:
    simple_commands = ["process", "ignore"]  # "@@CMD\n" or "@@CMD@@..."
    ## oid = object id; onum = object number
    compound_commands = ["meta", "link", "new", "oid", "onum"]  # "@@CMD:value@@"
    result: tuple[str, Command, str]    ## type, command, other
    if phrase.lower() in simple_commands:
        result = ("V", Command(phrase.lower(), []), "")
    elif re.search(r"^\d[\d& /,]*$", phrase):
        ## to catch alternative short form @@1&2 (= @@NEW:1&2)
        result = ("VO", Command("new", [phrase]), "")
    else:
        verb_object = phrase.split(":")
        if len(verb_object) == 1:
            result = ("unknown", Command("", []), "")
        else:
            verb, raw_object = verb_object
            verb = verb.lower()
            if verb in compound_commands:
                result = ("VO", Command(verb, raw_object.split("=")), "")
            else:
                result = ("none", Command("", []), phrase)
    return result


def process_verb(cmd: Command, content: Content) -> Content:
    match cmd.verb:
        case "ignore":
            content.currently_ignoring = True
        case "process":
            content.currently_ignoring = False
        case "unknown":
            msg = f"unknown verb ({cmd.object_list})"
            logging.warning(msg)
    overview.count[cmd.verb] += 1
    return content


def process_verb_object(cmd: Command, content: Content, concordance) -> Content:
    match cmd.verb:
        case "new":
            if (key_count := len(content.current_section_keys)) > 1:
                overview.count["EXTRA_sections"] += key_count - 1
            new_section_keys = [
                el for el in re.split(r"[\s&,]", cmd.object_list[0]) if el
            ]
            content.start_new_section(new_section_keys)
            content.currently_ignoring = True  ## ignore up to "process" in Penny
        case "meta":
            if cmd.object_list[0].lower() == "pub_date":
                content.pub_date = cmd.object_list[1]
            else:
                msg = f"Unknown META value: {cmd.object_list}"
                logging.warning(msg)
        case "link":
            text = cmd.object_list[0]
            content.add_to_line(text)
            # content.line = text
            # currently, links are ignored, but see this possibility:
            # number = re.sub(r"[^\d]", "", text)
            # if number:
            #     ref = concordance.get(number, ["", ""])[1]
            #     content.line = f" [{ref}]" if ref else ""
        # TODO: work out how to save these per section; use in preference to concordance, if available
        case "oid":
            content.current_info["object_id"] = cmd.object_list[0]
        case "onum":
            content.current_info["object_num"] = cmd.object_list[0]
        case _:
            msg = f"The command '{cmd.verb}' is unknown."
            logging.warning(msg)
    overview.count[cmd.verb] += 1
    return content


def create_shared_description_message(
    current_sections: list[str], concordance: dict[str, list[str]]
) -> str:
    shared_blurb = " & ".join(
        (
            f"{section} ({concordance.get(section, ("","UNKNOWN"))[1]})"
            for section in current_sections
        )
    )
    message = (
        f"[Description shared between {len(current_sections)} items: {shared_blurb}]"
    )
    logging.info(message)
    return message


def prepare_for_csv(
    content: Content,
    concordance: dict[str, list[str]],
    import_identifier: str,
    filter_by_id=[],
) -> list[ExcelRow]:
    headings = (
        "ID",       # Museum+ id of the item described
        "Import identifier",  # name given to this batch operation
        "Audience", # Always 'public'
        "Date",     # DD/MM/YYYY
        "Notes",
        "Purpose",
        "Sort",     # Always '100' in case of multiple entries
        "Source",
        "Status",   # Always '05 Published'
        "Text",
        "Title / Ref. No.",
        "Type",     # Always catalogue text
        "Language",
    )
    output: list[ExcelRow] = []
    object_id: str
    audience = "public"
    purpose = ""
    title = ""
    notes = ""
    source = ""
    status = "05 Published"
    _type = "catalogue text"
    language = "en"
    output.append(headings)
    # for num, lines in content.processed_text.items():
    for num, lines in content.processed_sections.items():
        object_id, catalogue_number = concordance.get(num, None)
        if filter_by_id and object_id in filter_by_id:
            continue
        if not object_id:
            logging.critical(f"No object id found in concordance for record {num}.")
            overview.missing["from_concordance"].append(int(num))
            # overview["from_concordance"].append(num)
        else:
            overview.count["records_output"] += 1
            _sort = "100"
            text = "\n\n".join(lines)
            output.append(
                (
                    object_id,
                    import_identifier,
                    audience,
                    content.pub_date,
                    notes,
                    purpose,
                    _sort,
                    source,
                    status,
                    text,
                    title,
                    _type,
                    language,
                )
            )
    return output


def overview_report() -> str:
    # print(overview.count)
    # print(overview.missing)
    report = "*" * 70 + "\n"
    report += "\t** Overview of commands performed:\n"
    report += f"\t** {", ".join([f"{key}={val}" for key, val in overview.count.items()])}\n"
    report += f"\t** Each section included a 'process' statement: {overview.count["new"] == overview.count["process"]}\n"
    report += f"\t** Total number of sections processed, including ones sharing same description = {overview.count["new"] + overview.count["EXTRA_sections"]}\n"
    report += f"\t** Total number of sections output to csv = {overview.count["records_output"]}\n"
    missing = overview.missing["from_concordance"]
    missing_details = (
        # f" (record no.{"s" if len(missing) > 1 else ""}: {', '.join([str(el) for el in missing])})"
        f" (record no.{plural_s(len(missing))}: {', '.join([str(el) for el in missing])})"
        if len(missing)
        else ""
    )
    report += f"\t** Total number of sections without links in concordance: {len(missing)}{missing_details}\n"
    report += "\t" + ("*" * 73)
    return report

def plural_s(count):
    return "s" if count > 1 else ""


def main() -> None:
    text_dir = Path("text_files")
    csv_dir = Path("csv_files")
    concordance_file = Path("penny.concordance.xlsx")
    if concordance_file.with_suffix(".json").exists():
        concordance = load_json_file(concordance_file.with_suffix(".json"))
    else:
        concordance = make_concordance_from_excel(concordance_file)
        export_dict_to_json(concordance, concordance_file)
    if not csv_dir.exists():
        csv_dir.mkdir()
    # quit()
    for source_file in text_dir.glob("*.txt"):
        overview.count.clear()
        overview.missing.clear()
        destination_file = csv_dir / f"{source_file.stem}.csv"
        batch_name = source_file.stem
        logging.info(
            f"Reading from {source_file.name} and writing to {destination_file.name}..."
        )
        raw_lines: list[str] = read_lines(source_file)
        content = process(raw_lines, concordance)
        del raw_lines
        csv_ready_text = prepare_for_csv(content, concordance, batch_name)
        # pprint(content)
        logging.info(overview_report())
        # export_to_csv(destination_file, csv_ready_text)
        export_to_excel(destination_file, csv_ready_text)


if __name__ == "__main__":
    main()
