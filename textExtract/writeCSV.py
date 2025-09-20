# import argparse
from pathlib import Path
from tools import shared
from csv import reader, writer
from openpyxl import load_workbook  # type: ignore[import]

from dataclasses import dataclass, field

# from typing import Iterable
import logging
import re
from collections import defaultdict
from copy import copy as shallow_copy

from pprint import pprint

logging.basicConfig(
    filename="textExtract.log",
    level=logging.INFO,
    # format='%(asctime)s %(levelname)s:%(message)s'
    format="%(levelname)s:%(message)s",
    filemode="w",
    encoding="utf-8",
)


@dataclass
class Overview:
    count: defaultdict[str, int] = field(default_factory=lambda: defaultdict(int))
    missing: defaultdict[str, list[int]] = field(
        default_factory=lambda: defaultdict(list)
    )

overview = Overview()

@dataclass
class Section:
    id: str
    oid: str
    onum: str
    paragraphs: list[str]


class Content:
    def __init__(self, pd: str, ps: dict, cl: list, line: str, csk: list, cinfo: dict, ci=True):
        self.pub_date: str = pd
        # self.processed_sections: dict[str, list[str]] = ps
        self.processed_sections: dict[str, Section] = ps
        # self.current_lines: list[str] = cl
        self.current = Section("", "", "", [])
        self.line: str = line
        self.current_section_keys: list[str] = csk
        # self.current_info: dict[str, str] = cinfo
        self.currently_ignoring: bool = ci

    def update_processed_sections(self):
        for key in self.current_section_keys:
            # self.processed_sections[key].paragraphs = self.current_lines
            self.processed_sections[key] = shallow_copy(self.current)
        # self.current_lines = []
        self.current = Section("", "", "", [])
        self.current_section_keys = []

    def start_new_section(self, new_section_keys: list[str]):
        self.update_processed_sections()
        self.current_section_keys = new_section_keys
        self.currently_ignoring = True

    def update_current_lines(self):
        if self.line:
            # self.current_lines.append(self.line)
            self.current.paragraphs.append(self.line)
            self.line = ""

    def add_to_line(self, part:str):
        # print(f">>>> {part[:50]}, {self.currently_ignoring}")
        if not self.currently_ignoring:
            self.line += part


@dataclass
class Command:
    verb: str
    object_list: list[str]


type ExcelRow = tuple[str, str, str, str, str, str, str, str, str, str, str, str, str]


# def argument_parser() -> tuple[Path, Path, Path]:
#     """Parse command line arguments."""
#     parser = argparse.ArgumentParser(description="Process some text files.")
#     parser.add_argument(
#         "-s", "--source", type=Path, default=Path("input.txt"), help="Source file path"
#     )
#     parser.add_argument(
#         "-o",
#         "--output",
#         type=Path,
#         default=Path("output.csv"),
#         help="Destination file path",
#     )
#     parser.add_argument(
#         "-c",
#         "--concordance",
#         type=Path,
#         default=Path("concordance.xlsx"),
#         help="Concordance file path",
#     )
#     args = parser.parse_args()
#     return (args.source, args.output, args.concordance)


# def read_lines(file_path: Path) -> list[str]:
#     """Read lines from a file and return them as a list."""
#     with file_path.open("r", encoding="utf-8") as file:
#         raw_lines = file.readlines()
#     # if ord(raw_lines[0][0]) == 65279:
#     #     raw_lines[0] = raw_lines[0][1:]
#     raw_lines[0] = remove_bom(raw_lines[0])
#     raw_lines = [line.strip() for line in raw_lines]
#     return raw_lines
#     # return file.readlines()


# def remove_bom(line: str) -> str:
#     # print(f">>> start character={ord(line[0])} ({ord(line[0]) == 65279})")
#     if ord(line[0]) == 65279:
#         line = line[2:]
#         logging.info(f"Removed BOM from start of file. <{line[:10]}>")
#         # return line[1:]
#     # if line.startswith("\xFF\xFE"):
#     #     logging.info("Removed BOM from start of file.")
#     #     return line[3:]
#     else:
#         logging.error("No BOM found.")
#     return line


# def write_lines(file_path: Path, lines: list[str]) -> None:
#     """Write a list of lines to a file."""
#     with file_path.open("w", encoding="utf-8") as file:
#         file.writelines(lines)


# def write_csv(file_path: Path, data: list) -> None:
#     """Write a list of lists to a CSV file."""
#     with file_path.open("w", encoding="utf-8", newline="") as file:
#         csv_writer = writer(file)
#         csv_writer.writerows(data)


# def read_csv(file_path: Path) -> list[list[str]]:
#     """Read a CSV file and return its content as a list of lists."""
#     with file_path.open("r", encoding="utf-8") as file:
#         csv_reader = reader(file)
#         return list(csv_reader)


# def extract_from_excel(excel_file_address: Path) -> list[list[str]]:
#     """
#     excel seems pretty random in how it assigns string/int/float, so...
#     this routine coerces everything into a string,
#     strips ".0" from misrecognised floats
#     & removes trailing spaces
#     """
#     excel_file_name: str = str(excel_file_address.resolve())
#     excel_sheet = load_workbook(filename=excel_file_name).active
#     sheet = []
#     if excel_sheet:
#         for excel_row in excel_sheet.iter_rows(min_row=2, values_only=True):
#             row = []
#             if not excel_row[0]:
#                 break
#             for col in excel_row:
#                 if col:
#                     data = str(col).strip()
#                     data = trim_mistaken_decimals(data)
#                 else:
#                     data = ""
#                 row.append(data)
#             sheet.append(row)
#     return sheet


# def trim_mistaken_decimals(string: str) -> str:
#     if string.endswith(".0"):
#         string = string[:-2]
#     return string


def make_concordance(excel_file_path: str) -> dict[str, list[str]]:
    """
    concordance layout with example data:
    row 0*  1*                   2                           3                   4                       5 (key)
    ObjID	ObjectNumber	    ObjectNumberSorted	        ReferenceNumber	    ObjectTitle	            PW:litCatNo
    742939	WA1947.191.176.1	WA1947.00191.00176.00001	Penny (1992) 1	    Flagellator of Christ	1

    PW:litCatNo (column 5) is the number given in the publication, so it is the key to the object id
    """
    raw = shared.extract_from_excel(Path(excel_file_path))
    concordance = normalise_concordance(raw)
    return concordance


def normalise_concordance(raw: list[list[str]]) -> dict[str, list[str]]:
    concordance: dict[str, list[str]] = {}
    for row in raw:
        object_id = row[0]
        object_num = row[1]
        key = row[5]
        if key.isnumeric():
            concordance[key] = [object_id, object_num]
    return concordance


def process(raw_lines: list[str], concordance: dict[str, list[str]]) -> Content:
    """
    This will replace group_lines(), get_command(), get_inline_commands() & parse_inline_command()
    This sorts the lines (paragraphs) into sections
    filters out lines that aren't needed
    and expands any inline commands
    """
    content: Content = Content("", {}, [], "", [],{})
    for raw_line in raw_lines:
        if not raw_line:
            continue
        phrases = chunk_by_command(raw_line)
        for _type, cmd, other in phrases:
            match _type:
                case "none":
                    content.add_to_line(other)
                case "V":
                    content = process_verb(cmd, content)
                case "VO":
                    content = process_verb_object(cmd, content, concordance)
                case "unknown":
                    msg = f">>>> UNKNOWN command: {cmd}"
                    logging.error(msg)
        content.update_current_lines()
    # if content.current_lines:
    if content.current.paragraphs:
        content.update_processed_sections()
    if content.pub_date:
        logging.info(f"Pub date: {content.pub_date}")
    else:
        logging.critical("No publication date found!")
    return content


def chunk_by_command(line: str) -> list[tuple[str, Command, str]]:
    open_cmd = False
    output = []
    pre_parsed_line = [el for el in re.split(r"(@@)", line) if el]
    ## re group preserves "@@": allows it to be matched with command to distinguish "PROCESS" (text) vs "@@PROCESS" (cmd)
    for phrase in pre_parsed_line:
        if phrase == "@@":
            open_cmd = not open_cmd
            continue
        if open_cmd:
            _type, cmd, other = parse_command(phrase)
        else:
            _type = "none"
            cmd = Command("", [])
            other = phrase
            # print(f">>>> {phrase[:50]}")
        output.append((_type, cmd, other))
    return output


def parse_command(phrase: str) -> tuple[str, Command, str]:
    simple_commands = ["process", "ignore"]  # "@@CMD\n" or "@@CMD@@..."
    ## oid = object id; onum = object number
    compound_commands = ["meta", "link", "new", "oid", "onum"]  # "@@CMD:value@@"
    result: tuple[str, Command, str]    ## type, command, other
    if phrase.lower() in simple_commands:
        result = ("V", Command(phrase.lower(), []), "")
    elif re.search(r"^\d[\d& /,]*$", phrase):
        ## to catch alternative short form @@1&2 (= @@NEW:1&2)
        result = ("VO", Command("new", [phrase]), "")
    else:
        verb_object = phrase.split(":")
        if len(verb_object) == 1:
            result = ("unknown", Command("", []), "")
        else:
            verb, raw_object = verb_object
            verb = verb.lower()
            if verb in compound_commands:
                result = ("VO", Command(verb, raw_object.split("=")), "")
            else:
                result = ("none", Command("", []), phrase)
    return result


def process_verb(cmd: Command, content: Content) -> Content:
    match cmd.verb:
        case "ignore":
            content.currently_ignoring = True
        case "process":
            content.currently_ignoring = False
        case "unknown":
            msg = f"unknown verb ({cmd.object_list})"
            logging.warning(msg)
    overview.count[cmd.verb] += 1
    return content


def process_verb_object(cmd: Command, content: Content, concordance) -> Content:
    match cmd.verb:
        case "new":
            if (key_count := len(content.current_section_keys)) > 1:
                overview.count["EXTRA_sections"] += key_count - 1
            new_section_keys = [
                el for el in re.split(r"[\s&,]", cmd.object_list[0]) if el
            ]
            content.start_new_section(new_section_keys)
            content.currently_ignoring = True  ## ignore up to "process" in Penny
        case "meta":
            if cmd.object_list[0].lower() == "pub_date":
                content.pub_date = cmd.object_list[1]
            else:
                msg = f"Unknown META value: {cmd.object_list}"
                logging.warning(msg)
        case "link":
            text = cmd.object_list[0]
            content.add_to_line(text)
            # content.line = text
            # currently, links are ignored, but see this possibility:
            # number = re.sub(r"[^\d]", "", text)
            # if number:
            #     ref = concordance.get(number, ["", ""])[1]
            #     content.line = f" [{ref}]" if ref else ""
        case "oid":
            content.current.oid = cmd.object_list[0]
        case "onum":
            content.current.onum = cmd.object_list[0]
        case _:
            msg = f"The command '{cmd.verb}' is unknown."
            logging.warning(msg)
    overview.count[cmd.verb] += 1
    return content


def create_shared_description_message(
    current_sections: list[str], concordance: dict[str, list[str]]
) -> str:
    shared_blurb = " & ".join(
        (
            f"{section} ({concordance.get(section, ("","UNKNOWN"))[1]})"
            for section in current_sections
        )
    )
    message = (
        f"[Description shared between {len(current_sections)} items: {shared_blurb}]"
    )
    logging.info(message)
    return message


def apply_concordance(content:Content, concordance: dict[str, list[str]]) -> Content:
    for num, section in content.processed_sections.items():
        if not section.oid:
            section.oid, section.onum = concordance.get(num, ["",""])
        if not section.oid:
            logging.warning(f"No object id found in concordance for record {num}.")
            overview.missing["from_concordance"].append(int(num))
    return content


def prepare_for_csv(
    content: Content,
    import_identifier: str,
) -> list[ExcelRow]:
    headings = (
        "ID",  # Museum+ id of the item described
        "Import identifier",  # name given to this batch operation
        "Type",  # Always catalogue text
        "Sort",  # Always '100' in case of multiple entries
        "Purpose",
        "Audience",  # Always 'public'
        "Status",  # Always '05 Published'
        "Language",
        "Published Date",  # DD/MM/YYYY
        "Title / Ref. No.",
        "Text",
        "Source",
        "Notes",
    )
    output: list[ExcelRow] = []
    output.append(headings)
    for num, section in content.processed_sections.items():
        if section.oid:
            overview.count["records_output"] += 1
            output.append(
                (
                    section.oid,
                    import_identifier,
                    "catalogue text",
                    "100",
                    "",
                    "public",
                    "05 Published",
                    "en",
                    content.pub_date,
                    "",
                    "\n\n".join(section.paragraphs),
                    "",
                    "",
                )
            )
        else:
            logging.critical(f"Record {num} excluded from csv output as it lacks an object ID.")
            # overview.missing["from_concordance"].append(int(num))
    return (output)


def overview_report() -> str:
    report = "*" * 70 + "\n"
    report += "\t** Overview of commands performed:\n"
    report += f"\t** {", ".join([f"{key}={val}" for key, val in overview.count.items()])}\n"
    report += f"\t** Each section included a 'process' statement: {overview.count["new"] == overview.count["process"]}\n"
    report += f"\t** Total number of sections processed, including ones sharing same description = {overview.count["new"] + overview.count["EXTRA_sections"]}\n"
    report += f"\t** Total number of sections output to csv = {overview.count["records_output"]}\n"
    missing = overview.missing["from_concordance"]
    missing_details = (
        # f" (record no.{"s" if len(missing) > 1 else ""}: {', '.join([str(el) for el in missing])})"
        f" (record no.{plural_s(len(missing))}: {', '.join([str(el) for el in missing])})"
        if len(missing)
        else ""
    )
    report += f"\t** Total number of sections without links in concordance: {len(missing)}{missing_details}\n"
    report += "\t" + ("*" * 73)
    return report


def plural_s(count):
    return "s" if count > 1 else ""


def update_text(outfile: Path, data: list[str], content:Content) -> None:
    """
    An attempt to update the file to add in the oid & onum
    """
    with outfile.open("w", encoding="utf-8") as f:
        ## print only first of a run of blank lines
        prev_was_blank = True
        for i, line in enumerate(data):
            if not line and prev_was_blank:
                continue
            elif not line:
                f.write(line + "\n")
                prev_was_blank = True
                continue
            else:
                prev_was_blank = False
            if re.match(r"@@\d", line):
                keys = line[2:].strip().split("&")
                oid, onum = [], []
                for key in keys:
                    oid.append(content.processed_sections[key].oid)
                    onum.append(content.processed_sections[key].onum)
                f.write("\n")
                f.write(f"@@NEW:{'&'.join(keys)}\n")
                f.write(f"@@OID:{'&'.join(oid)}\n")
                f.write(f"@@ONUM:{'&'.join(onum)}\n")
            else:
                f.write(line + "\n")


def main() -> None:
    text_dir = Path("text_files")
    csv_dir = Path("csv_files")
    if not csv_dir.exists():
        csv_dir.mkdir()
    update_dir = Path("updates")
    if not update_dir.exists():
        update_dir.mkdir()
    concordance = make_concordance("penny.concordance.xlsx")
    for source_file in text_dir.glob("*.txt"):
        overview.count.clear()
        overview.missing.clear()
        destination_file = csv_dir / f"{source_file.stem}.csv"
        updated_file = update_dir / f"{source_file.stem}.updated.txt"
        batch_name = source_file.stem
        logging.info(
            f"Reading from {source_file.name} and writing to {destination_file.name}..."
        )
        raw_lines: list[str] = shared.read_lines(source_file)
        content = process(raw_lines, concordance)
        content = apply_concordance(content, concordance)
        csv_ready_text = prepare_for_csv(content, batch_name)
        # update_text(update_dir / f"{source_file.stem}.updated.txt", raw_lines, content)
        update_text(updated_file, raw_lines, content)
        del raw_lines
        logging.info(overview_report())
        shared.write_csv(destination_file, csv_ready_text)


if __name__ == "__main__":
    main()
