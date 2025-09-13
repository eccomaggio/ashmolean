from pathlib import Path
from csv import reader, writer
import argparse
from openpyxl import load_workbook  # type: ignore[import]

# import datetime
# import pytz
from enum import Enum, auto
from dataclasses import dataclass, field

# from typing import Iterable
import logging
import re
from collections import defaultdict

# from pprint import pprint

# overview: defaultdict[str, int] = defaultdict(int)
# # overview['missing_from_concordance'] = []
# overview_missing: defaultdict[str, list] = defaultdict(lambda: [])

logging.basicConfig(
    filename="textExtract.log",
    level=logging.INFO,
    # format='%(asctime)s %(levelname)s:%(message)s'
    format="%(levelname)s:%(message)s",
    filemode="w",
    encoding="utf-8",
)


@dataclass
class Overview:
    count: defaultdict[str, int] = field(default_factory=lambda: defaultdict(int))
    missing: defaultdict[str, list[int]] = field(
        default_factory=lambda: defaultdict(list)
    )


overview = Overview()


class Content:
    def __init__(self, pd: str, ps: dict, cl: list, line: str, csk: list, ci=True):
        self.pub_date: str = pd
        self.processed_sections: dict[str, list[str]] = ps
        self.current_lines: list[str] = cl
        self.line: str = line
        self.current_section_keys: list[str] = csk
        self.currently_ignoring: bool = ci

    def update_processed_sections(self):
        for key in self.current_section_keys:
            self.processed_sections[key] = self.current_lines
        self.current_lines = []
        self.current_section_keys = []

    def start_new_section(self, new_section_keys: list[str]):
        self.update_processed_sections()
        self.current_section_keys = new_section_keys
        self.currently_ignoring = True

    def update_current_lines(self):
        if self.line:
            self.current_lines.append(self.line)
            self.line = ""

    def add_to_line(self, part:str):
        # print(f">>>> {part[:50]}, {self.currently_ignoring}")
        if not self.currently_ignoring:
            self.line += part


# class Instruction(Enum):
#     NONE = auto()
#     NEW_SECTION = auto()
#     PROCESS = auto()
#     IGNORE = auto()
#     META = auto()
#     UNKNOWN = auto()


# @dataclass
# class Command:
#     order: Instruction
#     details: list[str]
@dataclass
class Command:
    verb: str
    object_list: list[str]


type ExcelRow = tuple[str, str, str, str, str, str, str, str, str, str, str, str, str]


def argument_parser() -> tuple[Path, Path, Path]:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Process some text files.")
    parser.add_argument(
        "-s", "--source", type=Path, default=Path("input.txt"), help="Source file path"
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("output.csv"),
        help="Destination file path",
    )
    parser.add_argument(
        "-c",
        "--concordance",
        type=Path,
        default=Path("concordance.xlsx"),
        help="Concordance file path",
    )
    args = parser.parse_args()
    return (args.source, args.output, args.concordance)


def read_lines(file_path: Path) -> list[str]:
    """Read lines from a file and return them as a list."""
    with file_path.open("r", encoding="utf-8") as file:
        raw_lines = file.readlines()
    # if ord(raw_lines[0][0]) == 65279:
    #     raw_lines[0] = raw_lines[0][1:]
    raw_lines[0] = remove_bom(raw_lines[0])
    raw_lines = [line.strip() for line in raw_lines]
    return raw_lines
    # return file.readlines()


def remove_bom(line: str) -> str:
    # print(f">>> start character={ord(line[0])} ({ord(line[0]) == 65279})")
    if ord(line[0]) == 65279:
        line = line[2:]
        logging.info(f"Removed BOM from start of file. <{line[:10]}>")
        # return line[1:]
    # if line.startswith("\xFF\xFE"):
    #     logging.info("Removed BOM from start of file.")
    #     return line[3:]
    else:
        logging.error("No BOM found.")
    return line


def write_lines(file_path: Path, lines: list[str]) -> None:
    """Write a list of lines to a file."""
    with file_path.open("w", encoding="utf-8") as file:
        file.writelines(lines)


def write_csv(file_path: Path, data: list) -> None:
    """Write a list of lists to a CSV file."""
    with file_path.open("w", encoding="utf-8", newline="") as file:
        csv_writer = writer(file)
        csv_writer.writerows(data)


# def read_csv(file_path: Path) -> list[str]:
def read_csv(file_path: Path) -> list[list[str]]:
    """Read a CSV file and return its content as a list of lists."""
    with file_path.open("r", encoding="utf-8") as file:
        csv_reader = reader(file)
        return list(csv_reader)


def extract_from_excel(excel_file_address: Path) -> list[list[str]]:
    """
    excel seems pretty random in how it assigns string/int/float, so...
    this routine coerces everything into a string,
    strips ".0" from misrecognised floats
    & removes trailing spaces
    """
    excel_file_name: str = str(excel_file_address.resolve())
    excel_sheet = load_workbook(filename=excel_file_name).active
    sheet = []
    if excel_sheet:
        for excel_row in excel_sheet.iter_rows(min_row=2, values_only=True):
            row = []
            if not excel_row[0]:
                break
            for col in excel_row:
                if col:
                    data = str(col).strip()
                    data = trim_mistaken_decimals(data)
                else:
                    data = ""
                row.append(data)
            sheet.append(row)
    return sheet


def trim_mistaken_decimals(string: str) -> str:
    if string.endswith(".0"):
        string = string[:-2]
    return string


def make_concordance(excel_file_path: str) -> dict[str, list[str]]:
    """
    concordance layout with example data:
    row 0*  1*                   2                           3                   4                       5 (key)
    ObjID	ObjectNumber	    ObjectNumberSorted	        ReferenceNumber	    ObjectTitle	            PW:litCatNo
    742939	WA1947.191.176.1	WA1947.00191.00176.00001	Penny (1992) 1	    Flagellator of Christ	1

    PW:litCatNo (column 5) is the number given in the publication, so it is the key to the object id
    """
    raw = extract_from_excel(Path(excel_file_path))
    concordance = normalise_concordance(raw)
    return concordance


def normalise_concordance(raw: list[list[str]]) -> dict[str, list[str]]:
    concordance: dict[str, list[str]] = {}
    for row in raw:
        object_id = row[0]
        object_num = row[1]
        if row[5].isnumeric():
            concordance[row[5]] = [object_id, object_num]
    return concordance


def process(raw_lines: list[str], concordance: dict[str, list[str]]) -> Content:
    """
    This will replace group_lines(), get_command(), get_inline_commands() & parse_inline_command()
    This sorts the lines (paragraphs) into sections
    filters out lines that aren't needed
    and expands any inline commands
    """
    content: Content = Content("", {}, [], "", [])
    for raw_line in raw_lines:
        if not raw_line:
            continue
        phrases = chunk_by_command(raw_line)
        for _type, cmd, other in phrases:
            match _type:
                case "none":
                    content.add_to_line(other)
                case "V":
                    content = process_verb(cmd, content)
                case "VO":
                    content = process_verb_object(cmd, content, concordance)
                case "unknown":
                    msg = f">>>> UNKNOWN command: {cmd}"
                    logging.error(msg)
        content.update_current_lines()
    if content.current_lines:
        content.update_processed_sections()
    if content.pub_date:
        logging.info(f"Pub date: {content.pub_date}")
    else:
        logging.critical("No publication date found!")
    return content


def chunk_by_command(line: str) -> list[tuple[str, Command, str]]:
    open_cmd = False
    output = []
    pre_parsed_line = [el for el in re.split(r"(@@)", line) if el]
    ## re group preserves "@@": allows it to be matched with command to distinguish "PROCESS" (text) vs "@@PROCESS" (cmd)
    for phrase in pre_parsed_line:
        if phrase == "@@":
            open_cmd = not open_cmd
            continue
        if open_cmd:
            _type, cmd, other = parse_command(phrase)
        else:
            _type = "none"
            cmd = Command("", [])
            other = phrase
            # print(f">>>> {phrase[:50]}")
        output.append((_type, cmd, other))
    return output


def parse_command(phrase: str) -> tuple[str, Command, str]:
    simple_commands = ["process", "ignore"]  # "@@CMD\n" or "@@CMD@@..."
    compound_commands = ["meta", "link", "new"]  # "@@CMD:value@@"
    result: tuple[str, Command, str]
    # _type = ""
    # cmd = Command("", [])
    # other = ""
    if phrase.lower() in simple_commands:
        # cmd.verb = phrase.lower()
        # _type = "V"
        result = ("V", Command(phrase.lower(), []), "")
    else:
        verb_object = phrase.split(":")
        if len(verb_object) == 1:
            # _type = "unknown"
            # other = phrase
            result = ("unknown", Command("", []), "")
        else:
            # cmd.verb, raw_object = verb_object
            # cmd.object_list = raw_object.split("=")
            # cmd.verb = cmd.verb.lower()
            # if cmd.verb in compound_commands:
            verb, raw_object = verb_object
            # object_list = raw_object.split("=")
            verb = verb.lower()
            if verb in compound_commands:
                # _type = "VO"
                result = ("VO", Command(verb, raw_object.split("=")), "")
            else:
                # _type = "none"
                # cmd.verb, cmd.object_list, other = "", [], phrase
                result = ("none", Command("", []), phrase)
    # return (_type, cmd, other)
    return result


def process_verb(cmd: Command, content: Content) -> Content:
    match cmd.verb:
        case "ignore":
            content.currently_ignoring = True
        case "process":
            content.currently_ignoring = False
        case "unknown":
            msg = f"unknown verb ({cmd.object_list})"
            logging.warning(msg)
            # print(msg)
    # content.line += f"**{cmd.verb}**"
    return content


def process_verb_object(cmd: Command, content: Content, concordance) -> Content:
    # if content.currently_ignoring:
    #     if cmd.verb == "new":
    #         content.currently_ignoring = False
    #     else:
            # return content
    match cmd.verb:
        case "new":
            if (key_count := len(content.current_section_keys)) > 1:
                overview.count["EXTRA_sections"] += key_count - 1
                # print(f"EXTRA_sections {key_count - 1}")
            new_section_keys = [
                el for el in re.split(r"[\s&,]", cmd.object_list[0]) if el
            ]
            content.start_new_section(new_section_keys)
            content.currently_ignoring = True  ## ignore up to "process" in Penny
        case "meta":
            if cmd.object_list[0].lower() == "pub_date":
                content.pub_date = cmd.object_list[1]
            else:
                msg = f"Unknown META value: {cmd.object_list}"
                logging.warning(msg)
        case "link":
            text = cmd.object_list[0]
            content.add_to_line(text)
            # content.line = text
            # currently, links are ignored, but see this possibility:
            # number = re.sub(r"[^\d]", "", text)
            # if number:
            #     ref = concordance.get(number, ["", ""])[1]
            #     content.line = f" [{ref}]" if ref else ""
        case _:
            # logging.warning(f"The command '{command}' is unknown.")
            msg = f"The command '{cmd.verb}' is unknown."
            logging.warning(msg)
            # content.line = str(cmd.object_list)

        # overview.count[command] += 1
    return content


def create_shared_description_message(
    current_sections: list[str], concordance: dict[str, list[str]]
) -> str:
    # shared_blurb = " & ". join((f"{section} ({concordance[section][1]})" for section in current_sections))
    shared_blurb = " & ".join(
        (
            f"{section} ({concordance.get(section, ("","UNKNOWN"))[1]})"
            for section in current_sections
        )
    )
    message = (
        f"[Description shared between {len(current_sections)} items: {shared_blurb}]"
    )
    logging.info(message)
    return message


def prepare_for_csv(
    content: Content,
    concordance: dict[str, list[str]],
    import_identifier: str,
) -> list[ExcelRow]:
    headings = (
        "ID",  # Museum+ id of the item described
        "Import identifier",  # name given to this batch operation
        "Type",  # Always catalogue text
        "Sort",  # Always '100' in case of multiple entries
        "Purpose",  # TODO: check this should be empty = 'purpose'
        "Audience",  # Always 'public'
        "Status",  # Always '05 Published'
        "Language",
        "Published Date",  # DD/MM/YYYY
        "Title / Ref. No.",  # TODO: check this should be empty = 'title/ref. no.'
        "Text",
        "Source",  # TODO: check this should be empty = 'source'
        "Notes",
    )
    output: list[ExcelRow] = []
    object_id: str
    audience = "public"
    purpose = ""
    title = ""
    notes = ""
    source = ""
    status = "05 Published"
    _type = "catalogue text"
    language = "en"
    output.append(headings)
    # for num, lines in content.processed_text.items():
    for num, lines in content.processed_sections.items():
        object_from_concordance = concordance.get(num, None)
        if not object_from_concordance:
            logging.critical(f"No object id found in concordance for record {num}.")
            overview.missing["from_concordance"].append(int(num))
            # overview["from_concordance"].append(num)
        else:
            overview.count["records_output"] += 1
            object_id = object_from_concordance[0]
            _sort = "100"
            text = "\n\n".join(lines)
            output.append(
                (
                    object_id,
                    import_identifier,
                    _type,
                    _sort,
                    purpose,
                    audience,
                    status,
                    language,
                    content.pub_date,
                    title,
                    text,
                    source,
                    notes,
                )
            )
    return output


def overview_report() -> str:
    report = "*" * 70 + "\n"
    report += "\t**Overview of commands performed:\n"
    report += f"\t**Each section included a 'process' statement: {overview.count["NEW_SECTION"] == overview.count["PROCESS"]}\n"
    report += f"\t**Total number of sections processed, including ones sharing same description = {overview.count["NEW_SECTION"] + overview.count["EXTRA_sections"]}\n"
    report += f"\t**Total number of sections output to csv = {overview.count["records_output"]}\n"
    missing = overview.missing["from_concordance"]
    missing_details = (
        f" (record no.{"s" if len(missing) > 1 else ""}: {', '.join([str(el) for el in missing])})"
        if len(missing)
        else ""
    )
    report += f"\t**Total number of sections without links in concordance: {len(missing)}{missing_details}\n"
    report += "\t" + ("*" * 73)
    return report


def main() -> None:
    text_dir = Path("text_files")
    csv_dir = Path("csv_files")
    if not csv_dir.exists():
        csv_dir.mkdir()
    concordance = make_concordance("penny.concordance.xlsx")
    for source_file in text_dir.glob("*.txt"):
        overview.count.clear()
        overview.missing.clear()
        destination_file = csv_dir / f"{source_file.stem}.csv"
        batch_name = source_file.stem
        # published_date = "01/01/1992"

        logging.info(
            f"Reading from {source_file.name} and writing to {destination_file.name}..."
        )
        raw_lines: list[str] = read_lines(source_file)
        # processed_text, published_date = process(raw_lines, concordance)
        content = process(raw_lines, concordance)
        del raw_lines
        csv_ready_text = prepare_for_csv(content, concordance, batch_name)
        # csv_ready_text = prepare_for_csv(
        #     processed_text, concordance, published_date, batch_name
        # )
        # del processed_text

        logging.info(overview_report())
        # logging.info(f"Processed {len(csv_ready_text) - 1} sections from {source_file.name}.\n\n" )
        write_csv(destination_file, csv_ready_text)


if __name__ == "__main__":
    main()
